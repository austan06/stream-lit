import os
import io
from datetime import datetime, timezone

import requests
from requests.auth import HTTPBasicAuth

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


# ----------------------------
# Utilities
# ----------------------------
def parse_datetime(date_str: str) -> str:
    if not date_str:
        return ""
    fmts = ["%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z"]
    for fmt in fmts:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            continue
    # fallback: try to return date portion if ISO-like
    try:
        return date_str.split("T")[0]
    except Exception:
        return ""


def parse_datetime_raw(date_str: str) -> datetime:
    fmts = ["%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z"]
    for fmt in fmts:
        try:
            return datetime.strptime(date_str, fmt)
        except Exception:
            continue
    # last resort - return now (caller should handle)
    return datetime.now(timezone.utc)


# ----------------------------
# Jira API Helpers
# ----------------------------
def fetch_issue_changelog(issue_key: str, base_url, auth, headers):
    url = f"{base_url}/rest/api/3/issue/{issue_key}/changelog"
    changelog = []
    start_at = 0
    while True:
        resp = requests.get(f"{url}?startAt={start_at}&maxResults=100", headers=headers, auth=auth, timeout=60)
        if resp.status_code == 200:
            data = resp.json()
            changelog.extend(data.get("values", []))
            if data.get("total", 0) > start_at + 100:
                start_at += 100
            else:
                break
        else:
            raise Exception(f"Failed to retrieve changelog for {issue_key}: {resp.text}")
    return changelog


def min_date(*dates):
    valid_dates = [d for d in dates if d]
    if not valid_dates:
        return ""
    # Parse to datetime for comparison
    parsed = []
    for d in valid_dates:
        try:
            parsed.append(datetime.strptime(d, "%Y-%m-%d"))
        except Exception:
            continue
    if not parsed:
        return ""
    return min(parsed).strftime("%Y-%m-%d")


def get_status_transition_dates(issue_key, base_url, auth, headers):
    changelog = fetch_issue_changelog(issue_key, base_url, auth, headers)
    status_dates = {}
    for change in changelog:
        created = change.get("created")
        if not created:
            continue
        for item in change.get("items", []):
            if item.get("field") == "status":
                status = item.get("toString")
                date = parse_datetime(created)
                if status and status not in status_dates:
                    status_dates[status] = date
    return status_dates


def get_blocked_days(issue_key, base_url, auth, headers) -> int:
    changelog = fetch_issue_changelog(issue_key, base_url, auth, headers)
    blocked_dates = []
    blocked_days = 0
    for change in changelog:
        created = change.get("created")
        if not created:
            continue
        for item in change.get("items", []):
            if item.get("field") == "status" and (
                    item.get("toString") == "Blocked" or item.get("fromString") == "Blocked"):
                blocked_dates.append(parse_datetime_raw(created))
    for i in range(0, len(blocked_dates), 2):
        start_date = blocked_dates[i]
        end_date = blocked_dates[i + 1] if i + 1 < len(blocked_dates) else datetime.now(timezone.utc)
        blocked_days += (end_date - start_date).days
    return blocked_days


def get_peer_review_unassigned_time(issue_key, base_url, auth, headers) -> float:
    if not issue_key or issue_key.strip() == "":
        return 0.0

    changelog = fetch_issue_changelog(issue_key, base_url, auth, headers)
    peer_review_periods = []
    assignee_changes = []

    for change in changelog:
        created = change.get("created")
        if not created:
            continue
        change_time = parse_datetime_raw(created)
        for item in change.get("items", []):
            field = item.get("field")
            if field == "status":
                if item.get("toString") == "Peer Review":
                    peer_review_periods.append({"start": change_time, "end": None})
                elif item.get("fromString") == "Peer Review" and peer_review_periods:
                    for period in reversed(peer_review_periods):
                        if period["end"] is None:
                            period["end"] = change_time
                            break
            elif field == "assignee":
                assignee_changes.append({
                    "time": change_time,
                    "from": item.get("fromString"),
                    "to": item.get("toString"),
                })

    now = datetime.now(timezone.utc)
    for period in peer_review_periods:
        if period["end"] is None:
            period["end"] = now

    total_unassigned = pd.Timedelta(0)
    for period in peer_review_periods:
        current_time = period["start"]
        assignee = None
        relevant_changes = [c for c in assignee_changes if period["start"] <= c["time"] <= period["end"]]
        relevant_changes.sort(key=lambda x: x["time"])
        for change in relevant_changes:
            if assignee is None:
                total_unassigned += (change["time"] - current_time)
            current_time = change["time"]
            assignee = change["to"]
        if assignee is None:
            total_unassigned += (period["end"] - current_time)

    return round(total_unassigned.total_seconds() / 86400, 2)


def fetch_issues(sprint, base_url, auth, headers, story_points_field, project_code_field):
    issues = []
    start_at = 0
    max_results = 100
    jql = f'Sprint = {sprint} AND status IN ("Testing", "Approved for Release", "Closed")'
    while True:
        SEARCH_URL = f"{base_url}/rest/api/3/search/jql"
        params = {
            "startAt": start_at,
            "maxResults": max_results
        }

        # Send JQL in the request body as JSON
        payload = {"jql": jql,
                   "fields": [
                       "summary", "key", "issuetype", "status", "created",
                       story_points_field, project_code_field, "priority",
                   ]
                   }

        resp = requests.post(SEARCH_URL, headers={**headers, "Content-Type": "application/json"}, auth=auth,
                             params=params, json=payload, timeout=90)
        if resp.status_code == 200:
            data = resp.json()
            issues.extend(data.get("issues", []))
            if start_at + max_results >= data.get("total", 0):
                break
            start_at += max_results
        else:
            raise Exception(f"Failed to retrieve issues: {resp.text}")
        print(f"Fetched {issues} issues so far...")
    return issues


# ----------------------------
# Excel Export (full headers + formulas)
# ----------------------------
def create_full_excel_with_formulas(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Flow Metrics"

    # Write dataframe to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define extra/final headers and place them after existing columns
    extra_headers = [
        "Region", "Cycle Time", "Lead Time", "In Progress Time", "PR Time",
        "Active PR Time", "PD Time", "Customer", "Ramp Time", "Flow Efficiency", "YearClean",
    ]
    start_col = ws.max_column + 1
    for idx, header in enumerate(extra_headers):
        ws.cell(row=1, column=start_col + idx, value=header).font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=19,
                value=f'=IFS(P{row}="Kraken","US",P{row}="TOS","US",P{row}="Argos","US",P{row}="Alchemy","US",TRUE,"International")')
        ws.cell(row=row, column=20,
                value=f"=IF(OR(O{row}=\"\",C{row}=\"\"),\"\",ROUND((DATEVALUE(O{row})-DATEVALUE(C{row})+1),2))")
        ws.cell(row=row, column=21,
                value=f"=IF(OR(O{row}=\"\",B{row}=\"\"),\"\",ROUND((DATEVALUE(O{row})-DATEVALUE(B{row})+1),2))")
        ws.cell(row=row, column=22,
                value=f"=IF(OR(D{row}=\"\",C{row}=\"\"),\"\",ROUND((DATEVALUE(D{row})-DATEVALUE(C{row})),2))")
        ws.cell(row=row, column=23,
                value=f'=IF(D{row}="",0,ROUND(IFS(E{row}<>"",DATEVALUE(E{row})-DATEVALUE(D{row}),F{row}<>"",DATEVALUE(F{row})-DATEVALUE(D{row}),G{row}<>"",DATEVALUE(G{row})-DATEVALUE(D{row}),H{row}<>"",DATEVALUE(H{row})-DATEVALUE(D{row})),0)+1)')
        ws.cell(row=row, column=24, value=f"=IF(OR(W{row}=\"\",N{row}=\"\"),\"\",W{row}-N{row})")
        ws.cell(row=row, column=25,
                value=f"=IF(E{row}=\"\",0,IF(O{row}=\"\",\"\",DATEVALUE(O{row})-DATEVALUE(E{row})+1))")
        ws.cell(row=row, column=26, value=f"=IF(A{row}=\"\",\"\",TEXTBEFORE(A{row},\"-\"))")
        ws.cell(row=row, column=27,
                value=f"=IF(OR(C{row}=\"\",B{row}=\"\"),\"\",ROUND(DATEVALUE(C{row})-DATEVALUE(B{row}),1))")
        ws.cell(row=row, column=28, value=f"=IF(OR(T{row}=\"\",U{row}=\"\",U{row}=0),0,T{row}/U{row})")
        ws.cell(row=row, column=28).number_format = '0.00%'
        ws.cell(row=row, column=29, value=f"=IF(R{row}=\"\",\"\",CLEAN(R{row}))")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="Jira Flow Metrics", layout="wide")
st.title("📊 Jira Sprint Metrics")

with st.sidebar:
    st.markdown("### Jira connection")
    sprint = st.text_input("Sprint ID (numeric)", "")
    jira_url = st.text_input("Jira Base URL", os.environ.get("JIRA_BASE_URL", "https://theplatform.jira.com"))
    email = st.text_input("Email", os.environ.get("JIRA_EMAIL", ""), type="default")
    api_token = st.text_input("API Token", os.environ.get("JIRA_API_TOKEN", ""), type="password")
    story_points_field = st.text_input("Story Points Field", "customfield_10013")
    project_code_field = st.text_input("Project Code Field", "customfield_14300")
    fetch_btn = st.button("Fetch & Build Excel")

if fetch_btn:
    if not (sprint and email and api_token and jira_url):
        st.error("Please provide Sprint ID, Jira Base URL, Email and API token.")
    else:
        st.info("Fetching Jira issues (may take a while for large sprints)...")
        try:
            auth = HTTPBasicAuth(email, api_token)
            headers = {"Accept": "application/json"}

            issues = fetch_issues(sprint, jira_url, auth, headers, story_points_field, project_code_field)

            # Build rows consistent with the full header set
            rows = []
            for issue in issues:
                fields = issue.get("fields", {}) or {}
                key = issue.get("key", "")
                created_raw = fields.get("created", "")
                # We'll map some statuses from changelog to dates
                status_dates = get_status_transition_dates(key, jira_url, auth, headers)

                # Fill status order dates in the expected columns
                backlog = status_dates.get("Backlog", "") or ""
                in_progress = status_dates.get("In Progress", "") or status_dates.get("InProgress", "") or ""
                peer_review = status_dates.get("Peer Review", "") or ""
                pending_deploy = status_dates.get("Pending Deployment", "") or ""
                testing = status_dates.get("Testing", "") or ""
                approved = status_dates.get("Approved for Release", "") or ""
                closed = status_dates.get("Closed", "") or ""

                issue_type = (fields.get("issuetype") or {}).get("name", "")
                story_points = fields.get(story_points_field, "")
                priority = (fields.get("priority") or {}).get("name", "")
                blocked_days = get_blocked_days(key, jira_url, auth, headers)
                pcode_field = fields.get(project_code_field, "")
                project_code_value = pcode_field.get("value", "") if isinstance(pcode_field, dict) else (
                            pcode_field or "")
                unassigned_peer_review_days = get_peer_review_unassigned_time(key, jira_url, auth, headers)
                end_date = min_date(testing, approved, closed) or ""
                team = fields.get("customfield_team", "") or ""
                sprint_field = fields.get("customfield_sprint", "") or ""
                year_val = ""
                try:
                    if end_date:
                        year_val = int(end_date.split("-")[0])
                except Exception:
                    year_val = ""

                # Compose canonical row per requested order:
                row = [
                    key,
                    backlog or (created_raw and parse_datetime(created_raw)) or "",
                    in_progress,
                    peer_review,
                    pending_deploy,
                    testing,
                    approved,
                    closed,
                    issue_type,
                    story_points,
                    priority,
                    blocked_days,
                    project_code_value,
                    unassigned_peer_review_days,
                    end_date,
                    team,
                    sprint_field,
                    year_val
                ]
                rows.append(row)

            columns = [
                "ID", "Backlog", "In Progress", "Peer Review", "Pending Deployment",
                "Testing", "Approved for Release", "Closed", "Issue Type", "Story Points",
                "Priority", "Blocked Days", "Project Code", "Unassigned Time in Peer Review (days)",
                "End Date", "Team", "Sprint", "Year"
            ]

            df = pd.DataFrame(rows, columns=columns)

            # Add calculated columns for preview (these will also be present in Excel via formulas)
            df["Region"] = df["Team"].apply(
                lambda x: "US" if x in ["Kraken", "TOS", "Argos", "Alchemy"] else "International")
            df["Cycle Time"] = (pd.to_datetime(df["End Date"], errors="coerce") - pd.to_datetime(df["In Progress"],
                                                                                                 errors="coerce")).dt.days + 1
            df["Lead Time"] = (pd.to_datetime(df["End Date"], errors="coerce") - pd.to_datetime(df["Backlog"],
                                                                                                errors="coerce")).dt.days + 1
            df["In Progress Time"] = (
                        pd.to_datetime(df["Peer Review"], errors="coerce") - pd.to_datetime(df["In Progress"],
                                                                                            errors="coerce")).dt.days
            df["PR Time"] = (
                        pd.to_datetime(df["Pending Deployment"], errors="coerce") - pd.to_datetime(df["Peer Review"],
                                                                                                   errors="coerce")).dt.days.fillna(
                0)
            df["Active PR Time"] = df["PR Time"] - pd.to_numeric(df["Unassigned Time in Peer Review (days)"],
                                                                 errors="coerce").fillna(0)
            df["PD Time"] = (pd.to_datetime(df["End Date"], errors="coerce") - pd.to_datetime(df["Pending Deployment"],
                                                                                              errors="coerce")).dt.days + 1
            df["Customer"] = df["ID"].apply(lambda x: x.split("-")[0] if isinstance(x, str) and "-" in x else "")
            df["Ramp Time"] = (pd.to_datetime(df["In Progress"], errors="coerce") - pd.to_datetime(df["Backlog"],
                                                                                                   errors="coerce")).dt.days
            df["Flow Efficiency"] = df.apply(
                lambda r: (r["Cycle Time"] / r["Lead Time"]) if pd.notna(r["Cycle Time"]) and pd.notna(
                    r["Lead Time"]) and r["Lead Time"] != 0 else 0, axis=1)
            df["YearClean"] = df["Year"].astype(str).str.strip().replace("nan", "")

            st.success(f"Fetched {len(df)} issues.")
            st.dataframe(df)

            # Excel with formulas
            excel_bytes = create_full_excel_with_formulas(df)
            st.download_button(
                "⬇️ Download Excel (with formulas)",
                data=excel_bytes,
                file_name="flow_metrics_with_formulas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error: {e}")

else:
    st.info("Enter Sprint ID and credentials in the sidebar and click *Fetch & Build Excel*.")
